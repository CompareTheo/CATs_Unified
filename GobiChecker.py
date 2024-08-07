#!/usr/bin/env python3
from tkinter import *
from tkinter import messagebox
from tkinter import filedialog
from tkinter import ttk
from tkinter import Label, PhotoImage
from PIL import Image, ImageTk
import configparser
import openpyxl
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

# local modules
import gobi
from alma import sru

# main program ################################################################
def resource_path(relative_path):
    """ Get the absolute path to the resource, works for dev and for PyInstaller """
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)

def main(*args):
    f_path = gui.openfile()
    if f_path == "":
        return
    
    # get row count
    row_count = sum(1 for line in open(f_path))

    # loop through and parse GOBI file
    gobi_file = open(f_path, 'r', encoding='utf-8')

    lines = gobi_file.readlines()

    headers = lines[0].strip().split("\t")


    for line in lines[1:]:
        
        # initiate Gobi order line object
        order = gobi.ParseLine(line, headers)
        
        # check for null lines and skip
        if order.line_is_null == True:
            continue
            
        # _____________________ PERFORM SRU SEARCHES _________________________#
        
        # generate SRU urls
        iz_isbn_query = sru.make_url(zone="IZ", sru_path=config.iz_SRU_path, 
                                       query=f"alma.isbn={order.isbn}")
        iz_title_query = sru.make_url(zone="IZ", sru_path=config.iz_SRU_path, 
                                       query=f'alma.title="{order.title_clean}"')
        iz_kw_query = sru.make_url(zone="IZ", sru_path=config.iz_SRU_path, 
                                       query=f'alma.all_for_ui all "{order.kw}"')
     
        urls = [
            iz_isbn_query,
            iz_title_query,
            iz_kw_query,
        ]
            
        (iz_isbn_query_resp, 
         iz_title_query_resp,
         iz_kw_query_resp,)= sru.searches(urls, 3)
        
        # create search objects
        iz_isbn = sru.parse(iz_isbn_query_resp, zone="IZ", 
                              inst_code=config.inst_code)
        iz_title = sru.parse(iz_title_query_resp, zone="IZ", 
                               inst_code=config.inst_code)
        iz_kw = sru.parse(iz_kw_query_resp, zone="IZ", 
                               inst_code=config.inst_code)
        
        #______________________ PARSE RESULTS ________________________________#        
        
        # IZ-ISBN search
        iz_isbn_recs_found = ""
        if iz_isbn.numberOfRecords > 0:
            iz_isbn_recs_found = "X"
        
        # IZ-Title search
        iz_title_recs_found = ""
        if iz_title.numberOfRecords > 0:
            iz_title_recs_found = "X"
                    
        # GOBI Purchase Option Lookup
        order_options_found = ""
        if order.purchase_option:
            order_options_found = "X"
        
        # _____________________ GENERATE OUTPUT ______________________________#
        results = ""
        tag = ""
            
        if iz_title_recs_found == "X":
            tag = "manual_lookup_title"
            results = "Matching Title Found"

        
        if iz_isbn_recs_found == "X":
            tag = "manual_lookup_isbn"
            results = "Matching ISBN Found"
            
            
        if order_options_found == "X":
            tag = "ok_to_order"
            results = "GOBI CATs Option(s) Available"

        if iz_title_recs_found != "X" and iz_isbn_recs_found != "X" and order_options_found != "X":
            tag = "no_purchase"
            results = "No CATs Qualifying Purchase Available"


        # insert results into gui
        gui.counter += 1
        increment = 100 / row_count
        gui.insert_text(gui.counter, (order.isbn, order.title.title(), order.author.title(), 
                          order.pub_year, order.binding, 
                          iz_isbn_recs_found, iz_title_recs_found, 
                          results), tag)
        gui.progress_bar.step(increment)
        continue
            
    # finish
    gui.progress_bar["value"] = 100
    gui.msgbox("Done.")
    gobi_file.close()


# Configurations ##############################################################
class configs:
    def __init__(self, configfile):
        self.configs = configs

        c_dict = configparser.ConfigParser()
        c_dict.read(configfile)
        
        self.version                 = c_dict['misc']['version']

        self.download_directory      = c_dict['misc']['download_directory'] \
                                           .replace('\\', '//')
        
        self.inst_code               = c_dict['SRU']['inst_code']
        self.iz_SRU_path             = c_dict['SRU']['iz_path']
        
        self.log_directory           = c_dict['log']['log_directory'] \
                                           .replace('\\', '//')

# Gui #########################################################################
class gui:
    def __init__(self, master):
        self.master = master
        
        master.title("CATs Purchase Search "+config.version)
        master.resizable(0, 0)
        master.minsize(width=1370, height=900)
        master.maxsize(width=1370, height=900)
        icon_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "images", "logo_small.ico")
        master.iconbitmap(icon_path)

        # Logo image
        logo_width = 1370
        logo_height = 378
        logo_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "images", "logo.jpg")
        original_logo = Image.open(logo_path)
        resized_logo = original_logo.resize((logo_width, logo_height))
        logo = ImageTk.PhotoImage(resized_logo)
        self.logo = Label(image=logo)
        self.logo.image = logo
        self.logo.pack()
        
        # counter
        self.counter = -1
        
        # frames
        self.top_frame = Frame(master)
        self.top_frame.pack(side='top', fill='both', expand=False)
        
        self.run_button = Button(self.top_frame, text="OPEN FILE AND RUN", 
                                                 font="Arial 14", 
                                                 command=main, 
                                                 relief="groove")
        self.run_button.pack(fill='both', side='left', expand=True)
        
        #self.save_img = PhotoImage(format = 'png', file= '.\images\save_icon.png')
        self.save_button = Button(self.top_frame, text="SAVE LOG", 
                                                  #image=self.save_img, 
                                                  font="Arial 14", 
                                                  command=self.save_log_xlsx, 
                                                  relief="groove")
        
        self.save_button.pack(fill='both', side='right', expand=False)
        
        self.mid_frame = Frame(master)
        self.mid_frame.pack(side='top', fill='both', expand=True)
        
        # tree view
        self.tree = ttk.Treeview(self.mid_frame)
        style = ttk.Style()
        style.theme_use('default')
        
        # binds
        self.tree.bind('<Control-c>', self.copy_isbn_keyboard)
        self.tree.bind("<Button-3>", self.popup)
        
        # tree columns
        self.tree['columns'] = ('isbn', 'title', 'author', 'pub_date', 
                                  'binding', 'iz_search_isbn', 
                                  'iz_search_title', 'gobipurchase', 'permalink')
                                  
        self.tree.heading('#0', text='#', anchor='w')
        self.tree.heading('isbn', text='ISBN', anchor="w")
        self.tree.heading('title', text='Title', anchor="w")
        self.tree.heading('author', text='Author', anchor="w")
        self.tree.heading('pub_date', text='Date', anchor="w")
        self.tree.heading('binding', text='Binding', anchor="w")
        self.tree.heading('iz_search_isbn', text='IZ-ISBN', anchor="w")
        self.tree.heading('iz_search_title', text='IZ-Title', anchor="w")
        self.tree.heading('gobipurchase', text="Results", anchor="w")
        self.tree.heading('permalink', text="Permalink", anchor="w")
        
        self.tree.column("#0", width=40)
        self.tree.column("isbn", width=105)
        self.tree.column("title", width=300)
        self.tree.column("author", width=85)
        self.tree.column("pub_date", width=50)
        self.tree.column("binding", width=50)
        self.tree.column("iz_search_isbn", width=50, anchor="center")
        self.tree.column("iz_search_title", width=45, anchor="center")
        self.tree.column("gobipurchase", width=300, anchor="center")
        self.tree.column("permalink", width=400, anchor="center")
        
        self.tree.pack(fill="both", expand=False, side="left")
        
        # scrollbar
        v_scrollbar = ttk.Scrollbar(self.mid_frame, orient="vertical", 
                                      command=self.tree.yview)
        v_scrollbar.place(x=1375, y=26, height=376)
        self.tree.configure(yscrollcommand=v_scrollbar.set)
       
        # tags
        self.tree.tag_configure('ok_to_order', background="#ecf0f1")
        self.tree.tag_configure('manual_lookup_isbn', background='#026873', foreground="#FFFFFF")
        self.tree.tag_configure('manual_lookup_title', background='#024959', foreground='#FFFFFF')
        self.tree.tag_configure('no_purchase', background='#8c3b4a', foreground="#FFFFFF")
       
        # progressbar
        style.configure("red.Horizontal.TProgressbar", foreground='red', 
                          background='#2381df')
        self.progress_bar = ttk.Progressbar(master, 
                              style="red.Horizontal.TProgressbar", 
                              orient='horizontal', mode='determinate')
        self.progress_bar.pack(fill="both", expand=False, side="top")
    
        
        self.popup_menu = Menu(master, tearoff=0)
        self.popup_menu.add_command(label="Copy ISBN",
                                    command=self.copy_isbn_mouse)
        self.popup_menu.add_command(label="Copy Title",
                                    command=self.copy_title_mouse)
        
    def popup(self, event):
        iid = self.tree.identify_row(event.y)
        if iid:
            self.tree.selection_set(iid)
            self.popup_menu.post(event.x_root, event.y_root)
        else:
            pass
        
    def copy_isbn_keyboard(self, event):
        curItem = self.tree.focus()
        item_dict = self.tree.item(curItem)
        isbn = item_dict['values'][0]
        root.clipboard_clear()
        root.clipboard_append(isbn)
        
    def copy_isbn_mouse(self):
        curItem = self.tree.focus()
        item_dict = self.tree.item(curItem)
        isbn = item_dict['values'][0]
        root.clipboard_clear()
        root.clipboard_append(isbn)
        
    def copy_title_mouse(self):
        curItem = self.tree.focus()
        item_dict = self.tree.item(curItem)
        title = item_dict['values'][1]
        root.clipboard_clear()
        root.clipboard_append(title)

    def msgbox(self, msg):
        messagebox.showinfo("Attention", msg)

    def openfile(self):
        self.filename =  filedialog.askopenfilename(initialdir = config.download_directory,
                                                    title = "Select file", 
                                                    filetypes = (("TXT files",
                                                                    "*.txt"),
                                                    ("all files","*.*")))
        return self.filename
        
    def insert_text(self, counter, msg, tags):
        self.tree.insert("", "end", text=counter, values=(msg), tags=tags)
        self.tree.yview_moveto(1)
        root.update()
        
    def save_log_csv(self):
        saved_log = open(config.log_directory+"gobi_checker_log.csv", 
                                               "w", 
                                               encoding="utf-8", 
                                               newline='')
        children = self.tree.get_children()
        for child in children:
            list = self.tree.item(child)["values"]
            w = csv.writer(saved_log, quoting=csv.QUOTE_ALL)
            w.writerow(list)
        saved_log.close()
        self.msgbox("LOG SAVED SUCCESFULLY.")
        
    def save_log_xlsx(self):
        wb = Workbook()

        # grab the active worksheet
        ws = wb.active
        
        # headers
        headers = ["ISBN", "Title", "Author", "Publisher", "Date", "Binding", 
                     "IZ-ISBN", "IZ-Title", "IZ-KW", "Results", "Intentional Duplicate" "Selector"]
        ws.append(headers)
        
        # rows
        children = self.tree.get_children()
        for child in children:
            list = self.tree.item(child)["values"]
            list[0] = f"'{list[0]}" # add ' to isbn string
            ws.append(list)
        
        # set column widths
        ws.column_dimensions['A'].width = "20"   # isbn
        ws.column_dimensions['B'].width = "75"   # title
        ws.column_dimensions['C'].width = "40"   # author
        ws.column_dimensions['D'].width = "20"   # publisher
        ws.column_dimensions['E'].width = "15"   # date
        ws.column_dimensions['F'].width = "15"   # date
        ws.column_dimensions['G'].width = "10"   # IZ-ISBN
        ws.column_dimensions['H'].width = "10"   # IS-Title
        ws.column_dimensions['I'].width = "10"   # IZ-KW
        ws.column_dimensions['M'].width = "75"  # Results
        ws.column_dimensions['N'].width = "10"  # Intentional Duplicate
        ws.column_dimensions['O'].width = "40"  # Selector
        
        # freeze header
        a = ws['A2']
        ws.freeze_panes = a
        
        # set header styles
        for rows in ws.iter_rows(min_row=1, max_row=1, min_col=1):
            for cell in rows:
                cell.fill = PatternFill(fgColor='FFD700', bgColor='FFFFFF', 
                                        fill_type='solid')
                cell.font = Font(size=14, 
                                 bold=True, 
                                 italic=True, 
                                 underline='single')
                cell.alignment = openpyxl.styles.Alignment(horizontal='general', 
                                                           vertical='center')

        # save the file
        wb.save(f"{config.log_directory}/gobi_checker_log.xlsx")
        self.msgbox("LOG SAVED SUCCESFULLY.")
        

# toplevel ####################################################################
config = configs('config.ini')

# gui
root = Tk()
gui = gui(root)
root.mainloop()